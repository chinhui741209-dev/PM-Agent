"""把 WBS 草稿匯出成 Excel（.xlsx）。使用 openpyxl。"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import WbsDraft, WbsNode

WBS_HEADERS = [
    "層級", "ID", "類型", "工作項", "負責單位", "初始階段",
    "開始日", "到期日", "估時(天)", "交付物", "依賴", "里程碑",
]
MS_HEADERS = ["里程碑", "日期", "對應交付物"]

_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TYPE_FILL = {
    "epic": PatternFill("solid", fgColor="EDE9FE"),
    "story": PatternFill("solid", fgColor="DBEAFE"),
    "task": PatternFill("solid", fgColor="D1FAE5"),
    "subtask": PatternFill("solid", fgColor="F3F4F6"),
}


def _flatten(nodes: list[WbsNode], depth: int = 1):
    for n in nodes:
        yield n, depth
        if n.children:
            yield from _flatten(n.children, depth + 1)


def _ms_name(draft: WbsDraft, milestone_id):
    for m in draft.milestones:
        if m.id == milestone_id:
            return m.name
    return ""


def build_wbs_workbook(draft: WbsDraft) -> bytes:
    wb = Workbook()

    # --- WBS 工作表 ---
    ws = wb.active
    ws.title = "WBS"
    ws.append(WBS_HEADERS)
    for c in ws[1]:
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(vertical="center")

    for node, depth in _flatten(draft.nodes):
        ws.append([
            depth,
            node.id,
            node.type,
            ("    " * (depth - 1)) + node.title,  # 以縮排呈現階層
            node.owner_unit or "",
            node.workflow_stage or "",
            node.start_date.isoformat() if node.start_date else "",
            node.due_date.isoformat() if node.due_date else "",
            node.estimate_days if node.estimate_days is not None else "",
            node.deliverable or "",
            ", ".join(node.dependencies) if node.dependencies else "",
            _ms_name(draft, node.milestone_id),
        ])
        # 依類型上色（第 3 欄類型 + 第 4 欄工作項）
        fill = _TYPE_FILL.get(node.type)
        if fill:
            ws.cell(row=ws.max_row, column=3).fill = fill

    ws.freeze_panes = "A2"
    widths = [6, 10, 9, 46, 12, 12, 12, 12, 9, 20, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- 里程碑工作表 ---
    ws2 = wb.create_sheet("里程碑")
    ws2.append(MS_HEADERS)
    for c in ws2[1]:
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    for m in draft.milestones:
        ws2.append([m.name, m.date.isoformat() if m.date else "", "、".join(m.deliverables)])
    for i, w in enumerate([28, 14, 40], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- 摘要列（資訊）---
    ws3 = wb.create_sheet("資訊", 0)
    ws3.append(["WBS 摘要"])
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.append(["草稿 ID", draft.id])
    ws3.append(["交付日", draft.delivery_date.isoformat() if draft.delivery_date else ""])
    ws3.append(["工作流", draft.workflow.name if draft.workflow else ""])
    ws3.append(["相關條件", draft.conditions or ""])
    ws3.append(["需求（節錄）", (draft.requirement_text or "")[:200]])
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
